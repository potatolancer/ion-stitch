import sys
from pathlib import Path
import openpyxl
import datetime
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
#from openpyxl.styles import colors, Font, Color, PatternFill, Border, Side, Alignment

NAME_ERROR = "NAME_ERROR"
CONF_DIR = "./conf/"
CONF_FILE = CONF_DIR + "conf.txt"
LOG_FILE = CONF_DIR + "log.txt"
PRIO_STR = "prio="
PRIO_SMALLER = PRIO_STR + "smaller"
PRIO_LARGER = PRIO_STR + "larger"
MATCH_STR = "match="
MATCH_SINGLE = MATCH_STR + "single"
MATCH_AVERAGE = MATCH_STR + "range"
LBL_GROUP = "GROUPS"
LBL_ID_FILTER = "ID Filters"
LBL_POS = "POSITIVE FILE: "
LBL_NEG = "NEGATIVE FILE: "
LBL_FINAL_DATA = "FINAL DATA COLUMN RANGE:"
LBL_EXCLUSION = "EXCLUDED COMPOUNDS:"
LBL_END = "EOF"
DEFAULT_GROUP_SIZE = 6
group_offset = 2
mtb_offset = 2
weeks_per_group = 6

USR_CONFIG = "./usr_conf.txt"

class ID_Filter():
	def __init__(self, id="NO_ID", filter_str="", prio=PRIO_LARGER, match_type=MATCH_SINGLE):
		self.id = id
		self.filter_str = filter_str
		self.prio = prio
		self.match_type = match_type
		self.col_range = []
		self.next = None

		self.set_col_range()

	def get_col_min(self):
		if (isinstance(self.col_range[0], str)):
			return get_col_letters_as_i(self.col_range[0])
		return self.col_range[0]

	def get_col_max(self):
		if (isinstance(self.col_range[1], str)):
			return get_col_letters_as_i(self.col_range[1])
		return self.col_range[1]

	def set_col_range(self):
		cur_col = ""
		for char in self.filter_str:
			if (char == ':') or (char == ' '):
				self.col_range.append(cur_col)
				cur_col = ""
			cur_col += char.strip()


class Conf():
	def __init__(self):
		self.ref_id = ""
		self.id_filter = None
		self.num_id_filters = 0
		self.group = {}
		self.final_col_range = [1, 1]
		self.exclusions = []
		self.logger = []
		self.mode = []
		self.pos = ""
		self.neg = ""
		self.stitch_filename = ""
		self.should_close = False

		setup_dir(self, CONF_DIR)
		f = open(LOG_FILE, 'w', encoding="utf-8")
		f.write(str(datetime.datetime.now()) + "\n")
		f.close()

		#self.load_conf(CONF_FILE)

	def load_conf(self, file):
		labels = [LBL_POS, LBL_NEG, LBL_GROUP, LBL_ID_FILTER, LBL_FINAL_DATA, LBL_EXCLUSION, LBL_END]

		if (not Path(file).is_file()):
			self.error(f"{file} not found")
			return

		f = open(file, 'r')
		last_lbl = ""
		last_id_filter = None
		ctr = 0
		for l in f:
			line = l.replace('\n', '')
			if (line in labels):
				last_lbl = line
			if ((line == last_lbl) or (l == '\n')):
				continue

			if (LBL_POS == last_lbl):
				self.pos = line
			elif (LBL_NEG == last_lbl):
				self.neg = line
			elif (LBL_GROUP == last_lbl):
				if (len(line) == 0):
					break
				wordlist = self.buf_wordlist_from_line(line)
				self.group[wordlist[0]] = wordlist[1]
			elif (LBL_ID_FILTER == last_lbl):
				if (len(line) == 0):
					break
				wordlist = self.buf_wordlist_from_line(line)

				id = wordlist[0]
				filter_str = ""
				for word in wordlist[1:-2]:
					filter_str += word + " "
				match_type = wordlist[-1]
				prio = wordlist[-2]
				self.add_id_filter(id, filter_str, prio, match_type)
				ctr += 1
			elif (LBL_FINAL_DATA == last_lbl):
				wordlist = self.buf_wordlist_from_line(line)
				self.final_col_range[0] = wordlist[0]
				self.final_col_range[1] = wordlist[1]
			#elif (LBL_EXCLUSION == last_lbl):
			#	self.exclusions.append(line)
		f.close()

		self.log("positive ion mode read from conf: " + self.pos)
		self.log("negative ion mode read from conf: " + self.neg)

	def print_id_filters(self):
		cur_id_filter = self.id_filter
		while (cur_id_filter != None):
			cur_id_filter = cur_id_filter.next

	def buf_wordlist_from_line(self, readline):
		cur_word = ''
		wordlist = []
		for i in range(0, len(readline)):
			char = readline[i]
			if ((char != ' ') and (char != '\n')):
				cur_word += char
			elif (cur_word != ''):
				wordlist.append(cur_word)
				cur_word = ''
		if (cur_word != ''):
				wordlist.append(cur_word)
		return wordlist

	def get_id_filter(self, i):
		ctr = 0
		cur_id_filter = self.id_filter
		while (cur_id_filter is not None):
			if (isinstance(i, int)):
				if (ctr == i):
					return cur_id_filter
			else:
				if (cur_id_filter.id == i):
					return cur_id_filter

			cur_id_filter = cur_id_filter.next
			ctr += 1

	def get_file(self, ion_mode):
		if (ion_mode == "pos"):
			return self.pos_file
		if (ion_mode == "neg"):
			return self.neg_file

	def rearrange_id_filter(self, id, new_i):
		id_filter = self.remove_id_filter(id)
		prev_id_filter = self.id_filter
		temp_id_filter = self.id_filter

		i = int(new_i)
		if (int(new_i) > self.num_id_filters):
			i = self.num_id_filters
		elif (int(new_i) <= 0):
			id_filter.next = self.id_filter
			self.id_filter = id_filter
			return

		ctr = 0
		while (temp_id_filter != None):
			if (ctr == i):
				break

			prev_id_filter = temp_id_filter
			temp_id_filter = temp_id_filter.next
			ctr += 1
		id_filter.next = temp_id_filter
		prev_id_filter.next = id_filter

	def rearrange_group(self, group, new_i):
		new_group_list = []
		new_group = {}

		corr_i = 0
		if (int(new_i) > len(self.group)):
			corr_i = len(self.group)
		elif (int(new_i) > 0):
			corr_i = int(new_i)
		for key in self.group:
			if (key != group):
				new_group_list.append(key)
		new_group_list.insert(corr_i, group)
		for i in range(0, len(new_group_list)):
			new_group[new_group_list[i]] = self.group[new_group_list[i]]

		self.group = new_group

	def remove_id_filter(self, id):
		temp_id_filter = self.id_filter
		prev_id_filter = self.id_filter

		if (self.id_filter.id == id):
			self.id_filter = self.id_filter.next

		while (temp_id_filter != None):
			if (temp_id_filter.id == id):
				break
			prev_id_filter = temp_id_filter
			temp_id_filter = temp_id_filter.next

		if (temp_id_filter == None):
			self.error("remove_id_filter(): error")
			return None

		prev_id_filter.next = temp_id_filter.next

		self.num_id_filters -= 1
		return temp_id_filter

	def add_id_filter(self, id, wordlist, prio, match_type):
		new_id_filter = ID_Filter(id, wordlist, prio, match_type)

		temp_id_filter = self.id_filter

		if (temp_id_filter != None):
			if (temp_id_filter.id == id):
				new_id_filter.next = self.id_filter.next
				self.id_filter = new_id_filter
				return
			while (temp_id_filter != None):
				if (temp_id_filter.next == None):
					temp_id_filter.next = new_id_filter
					break
				if (temp_id_filter.next.id == id):
					new_id_filter.next = temp_id_filter.next.next
					temp_id_filter.next = new_id_filter
					break
				temp_id_filter = temp_id_filter.next
		else:
			self.id_filter = new_id_filter

		self.num_id_filters += 1

	def add_group(self, groupname, size):
		self.group[groupname] = size

	def switch_prio(self, id):
		id_filter = self.get_id_filter(id)
		prio = id_filter.prio
		if (prio == PRIO_SMALLER):
			id_filter.prio = PRIO_LARGER
			return
		id_filter.prio = PRIO_SMALLER

	def log(self, str, space=True, nl=True, pr=False):
		new_str = str
		f = open(LOG_FILE, 'a', encoding="utf-8")
		if (space):
			new_str += ' '
		if (nl):
			new_str += '\n'
		f.write(new_str)
		f.close()

		if (pr):
			print(str)

	def save(self):
		self.log("saving conf")
		f = open(CONF_FILE, 'w')
		f.write(f"{LBL_POS}\n{self.pos}\n\n")
		f.write(f"{LBL_NEG}\n{self.neg}\n\n")
		f.write(LBL_GROUP + '\n')
		for g in self.group:
			f.write(f"{g} {self.group[g]} \n")
		f.write('\n')
		f.write(LBL_ID_FILTER + '\n')
		cur_id_filter = self.id_filter
		while (cur_id_filter != None):
			f.write(cur_id_filter.id + " ")
			for word in cur_id_filter.filter_str:
				f.write(word)
			f.write(f"{cur_id_filter.prio} ")
			f.write(f"{cur_id_filter.match_type}")
			f.write("\n")
			cur_id_filter = cur_id_filter.next
		f.write('\n' + LBL_FINAL_DATA + '\n')
		f.write(str(self.final_col_range[0]) + ' ' + str(self.final_col_range[1]) + '\n')
		f.write('\n' + LBL_EXCLUSION + '\n')
		for i in range(0, len(self.exclusions)):
			f.write(self.exclusions[i] + '\n')
		f.write('\n' + LBL_END + '\n')
		f.close()

	def error(self, error_msg):
		self.log(error_msg)
		self.should_close = True

	def check_error(self):
		msg = f"program terminating early...\nsee {LOG_FILE} for details"
		if (self.should_close):
			self.log(msg)
			print(msg)
			quit()

	def read_args(self, argv):
		n_args = len(argv)
		log_args = []
		self.log(str(n_args) + " arguments provided... ")

		for i in range(0, n_args):
			arg = argv[i]
			next_flag_i = 0
			for j in range(i+1, n_args):
				next_flag_i = j + 1
				if (argv[j][0] == '-'):
					next_flag_i = j
					break
			self.log(arg, nl=False)

			id = ""
			filter = []
			if (len(argv) > i+1):
				next_arg = argv[i+1]

			if (arg == "-h"):
				self.mode.append("help")
			elif (arg == "-R"):
				self.mode.append("Review")
			elif (arg == "-E"):
				self.mode.append("Export")
				self.stitch_filename = next_arg
			elif (arg == "--final_col_range"):
				if (len(argv[i+1:next_flag_i]) == 2):
					self.final_col_range = [next_arg, argv[i+2]]
			elif (arg == "--exclude"):
				self.exclusions.append(next_arg)
			elif (arg == "--include"):
				cnt = 0
				for e in self.exclusions:
					if (e == next_arg):
						self.exclusions.pop(i)
					cnt += 1
			elif (arg == "-pos"):
				self.pos = next_arg
			elif (arg == "-neg"):
				self.neg = next_arg
			elif (arg == "-mf"):
				if (self.num_id_filters > 1):
					self.rearrange_id_filter(next_arg, argv[i+2])
			elif (arg == "-mg"):
				if (len(self.group) > 1):
					self.rearrange_group(next_arg, argv[i+2])
			elif ("-af" in arg):
				prio = PRIO_LARGER
				if ('s' in arg):
					prio = PRIO_SMALLER

				match_type = MATCH_AVERAGE
				if ('w' in arg):
					match_type = MATCH_SINGLE

				filter_str = ""
				for word in argv[i+2:next_flag_i]:
					filter_str += word + " "
				self.add_id_filter(next_arg, filter_str, prio, match_type)
			elif (arg == "-ag"):
				self.add_group(next_arg, argv[i+2])
			elif (arg == "--prio_large"):
				self.get_id_filter(next_arg).prio = PRIO_LARGER
			elif (arg == "--prio_small"):
				self.get_id_filter(next_arg).prio = PRIO_SMALLER
			elif (arg == "--match_single"):
				self.get_id_filter(next_arg).match_type = MATCH_SINGLE
			elif (arg == "--match_average"):
				self.get_id_filter(next_arg).match_type = MATCH_AVERAGE
			elif (arg == "-rf"):
				for arg in argv[i+1:next_flag_i]:
					try:	self.remove_id_filter(arg)
					except:	continue
			elif (arg == "-rg"):
				for arg in argv[i+1:next_flag_i]:
					try:	self.group.pop(arg, None)
					except:	continue

		self.log("")

		return len(sys.argv)


	def print_log(self):
		f = open(LOG_FILE, 'r')
		for line in f:
			print(line.strip())
		f.close()


class xl_data():
	def __init__(self, filename):
		self.filename = filename
		self.data = []
		self.header = []
		self.header.append("NULL")
		self.mtb = []
		self.r_max = 0
		self.c_max = 0
		self.num_mtbs = 0

	def load_xl(self, conf):
		try:
			wb = openpyxl.load_workbook(filename = self.filename)
			conf.log("success")
		except:
			conf.error("fail")
			return

		ws = wb.worksheets[0]
		self.r_max = ws.max_row
		self.c_max = ws.max_column

		for r in range(1, self.r_max + 1):
			r_data = []
			mtb_name = NAME_ERROR

			for c in range(1, self.c_max + 1):
				cell = ws.cell(row=r, column=c)
				if (r == 1):
					self.header.append(cell.value)
					continue

				elif (self.header[c] == "Name"):
					mtb_name = cell.value

				r_data.append(cell.value)

			self.mtb.append(Metabolite(mtb_name, r_data))

		for i in range(1, len(self.mtb)):
			self.mtb[i].autoset_data(self.header, conf)

		self.mtb.pop(0)

	def append_mtb(self, mtb):
		self.mtb.append(mtb)
		self.num_mtbs += 1

	def log(self, conf):
		conf.log(self.filename)
		for i in range(0, len(self.mtb)):
			try:
				conf.log(f"{self.mtb[i].name + ':':<40}{self.mtb[i].summary_data}")
			except:
				continue

class Metabolite():
	def __init__(self, name, data):
		self.name = name
		self.data = data
		self.summary_data = {}
		self.group_data = {}
		self.selected_mode = ''
		self.detected_mode = ''

	def check_against_summary_data(self, othr, conf):
		for i in range(0, conf.num_id_filters):
			id_filter = conf.get_id_filter(i)
			id = id_filter.id
			data_self = float(self.summary_data[id])
			data_othr = float(othr.summary_data[id])

			try:
				if (data_self == data_othr):
					continue
			except:
				conf.log(f"{id} filter not found in {self.name}")
				continue

			prio = id_filter.prio
			if (prio == PRIO_SMALLER):
				if (data_self < data_othr):
					return True
				else:
					return False
			else:
				if (data_self > data_othr):
					return True
				else:
					return False
			
	def autoset_data(self, header, conf):
		for i in range(0, conf.num_id_filters):
			id_filter = conf.get_id_filter(i)
			data = []
			avg_data = []

			for i in range(0, len(header)):
				if (id_filter.match_type == MATCH_SINGLE):
						if (header[i].strip() == id_filter.filter_str.strip()):
							data.append(self.data[i-1])
				elif (id_filter.match_type == MATCH_AVERAGE):
						if ((i >= id_filter.get_col_min()) and (i <= id_filter.get_col_max())):
							avg_data.append(self.data[i-1])

			if (id_filter.match_type == MATCH_SINGLE):
				if (len(data) > 1):
					self.log(f"WARNING: Multiple instances of data found for filter id: {id_filter.id}\nThe first instance was selected: {data[0]}")
				try:
					self.summary_data[id_filter.id] = data[0]
				except:
					self.summary_data[id_filter.id] = ""
				continue
			elif (id_filter.match_type == MATCH_AVERAGE):
				try:
					self.summary_data[id_filter.id] = (sum(avg_data) / len(avg_data))
				except:
					self.summary_data[id_filter.id] = ""

		
		col_i = [
			int(get_col_letters_as_i(conf.final_col_range[0])),
			int(get_col_letters_as_i(conf.final_col_range[1])) ]
		ctr = 0
		for i in range(col_i[0], col_i[1] + 1):
			try:
				self.group_data["{0:03d}".format(ctr)] = self.data[i-1]
			except:
				pass
			ctr += 1

def get_col_letters_as_i(col_letters):
	letter_as_int = (ord(col_letters[0].upper()) - 64) * (26 ** (len(col_letters)-1))
	if (len(col_letters) == 1):
		return letter_as_int

	letter_as_int += get_col_letters_as_i(col_letters[1:])
	return letter_as_int

def get_i_as_col_letters(i):
	remainder = i / 26

	mod = i % 26
	if (mod == 0):
		remainder -= 1
		mod = 26

	col_letters = chr(mod + 64)

	if (remainder < 1):
		return col_letters

	col_letters = (get_i_as_col_letters(int(remainder))) + col_letters
	return col_letters

def sort_by_lowercase(list):
	temp_list = list.copy()
	association = {}
	new_list = []

	list_lower = list.copy()

	for i in range(0, len(list_lower)):
		list_lower[i] = list_lower[i].lower()
		association[list_lower[i]] = list[i]

	list_lower.sort()

	for i in range(0, len(list_lower)):
		new_list.append(association[list_lower[i]])

	return new_list


def stitch_modes(conf, mode_p, mode_n):
	unique_mtb = []
	for mtb in mode_p.mtb:
		if (mtb.name == NAME_ERROR):
			continue
		if (mtb.name not in unique_mtb):
			unique_mtb.append(mtb.name)

	for mtb in mode_n.mtb:
		if (mtb.name == NAME_ERROR):
			continue
		if (mtb.name not in unique_mtb):
			unique_mtb.append(mtb.name)

	unique_mtb = sort_by_lowercase(unique_mtb)

	mode_s = xl_data(conf.stitch_filename)
	mode_s.header = mode_p.header
	p_ctr = 0
	n_ctr = 0
	for i in range(0, len(unique_mtb)):
		mtb_p = None
		if (p_ctr < len(mode_p.mtb)):
			if (mode_p.mtb[p_ctr].name == unique_mtb[i]):
				mtb_p = mode_p.mtb[p_ctr]
				p_ctr += 1

		mtb_n = None
		if (n_ctr < len(mode_n.mtb)):
			if (mode_n.mtb[n_ctr].name == unique_mtb[i]):
				mtb_n = mode_n.mtb[n_ctr]
				n_ctr += 1

		mtb_s = None
		if (mtb_p == None):
			mtb_s = mtb_n
			mtb_s.detected_mode = 'N'
			mtb_s.selected_mode = 'N'
		elif (mtb_n == None):
			mtb_s = mtb_p
			mtb_s.detected_mode = 'P'
			mtb_s.selected_mode = 'P'
		else:
			mtb_s = mtb_n
			mtb_s.selected_mode = 'N'
			if (mtb_p.check_against_summary_data(mtb_n, conf)):
				mtb_s = mtb_p
				mtb_s.selected_mode = 'P'
			mtb_s.detected_mode = 'B'

		mtb_s.autoset_data(mode_s.header, conf)
		mode_s.append_mtb(mtb_s)

	return mode_s


def final_stitch(conf, ion_modes):
    if len(ion_modes) < 3:
        raise ValueError("Expected positive, negative, and base ion modes")

    mode_s = xl_data(conf.stitch_filename)
    mode_s.header = list(ion_modes[0].header)

    positive_mode = ion_modes[0]
    negative_mode = ion_modes[1]
    base_mode = ion_modes[2]

    p_ctr = 0
    n_ctr = 0

    for base_mtb in base_mode.mtb:
        detected_mode = ""
        mtb_s = base_mtb

        mtb_p = None
        if p_ctr < len(positive_mode.mtb):
            if positive_mode.mtb[p_ctr].name == base_mtb.name:
                mtb_p = positive_mode.mtb[p_ctr]
                detected_mode = "P"
                p_ctr += 1

        mtb_n = None
        if n_ctr < len(negative_mode.mtb):
            if negative_mode.mtb[n_ctr].name == base_mtb.name:
                mtb_n = negative_mode.mtb[n_ctr]
                detected_mode = "B" if detected_mode == "P" else "N"
                n_ctr += 1

        if base_mtb.selected_mode == "P":
            if mtb_p is not None and mtb_s.name == mtb_p.name:
                mtb_s = mtb_p
                mtb_s.selected_mode = "P"
            else:
                conf.log(f"{base_mtb.name}: positive value not found")
        else:
            if mtb_n is not None and mtb_s.name == mtb_n.name:
                mtb_s = mtb_n
                mtb_s.selected_mode = "N"
            else:
                conf.log(f"{base_mtb.name}: negative value not found")

        mtb_s.detected_mode = detected_mode
        mtb_s.autoset_data(mode_s.header, conf)
        mode_s.append_mtb(mtb_s)

    return mode_s


def setup_dir(conf, dir):
	try:
		Path.mkdir(dir)
		conf.log(f"directory created :: {dir}")
	except FileExistsError:
		conf.log(f"directory exists :: {dir}")
	except PermissionError:
		conf.log(f"directory access denied :: {dir}")
	except Exception as e:
		conf.log(f"directory {dir} :: {e}")


def count_dir_files(dir):
	path = Path(dir)
	cnt = sum(1 for entry in path.iterdir() if entry.is_file())
	return cnt


def create_xl(mode_p, mode_n, mode_s, conf):
	wb = Workbook()
	ws = wb.active
	ws.title = "Summary"
	dim_h = DimensionHolder(worksheet=ws)

	id_offset = 4
	header_names = ["Detected", "Selected", "Compound Name"]
	for i in range(0, conf.num_id_filters):
		id = conf.get_id_filter(i).id
		header_names.append("P: " + id)
		header_names.append("N: " + id)
	conf.log("\tappending headers")
	ws.append(header_names)

	conf.log("\tadding summary rows")
	p_ctr = 0
	n_ctr = 0
	for s_ctr in range(0, len(mode_s.mtb)):
		conf.check_error()
		cell_color = "000000"
		print(str(len(mode_p.mtb)) + " " + (str(p_ctr)))
		mtb_p = mode_p.mtb[p_ctr]
		mtb_n = mode_n.mtb[n_ctr]
		mtb_s = mode_s.mtb[s_ctr]

		a_in_c = False
		b_in_c = False
		
		if (p_ctr < len(mode_p.mtb) - 1):
			if (mtb_p.name == mtb_s.name):
				a_in_c = True
				p_ctr += 1
		if (n_ctr < len(mode_n.mtb) - 1):
			if (mtb_n.name == mtb_s.name):
				b_in_c = True
				n_ctr += 1

		cur_row = []
		cur_row.append(mtb_s.detected_mode)
		cur_row.append(mtb_s.selected_mode)
		cur_row.append(mtb_s.name)


		for i in range(0, conf.num_id_filters):
			id = conf.get_id_filter(i).id
			try:
				if (a_in_c):
					cur_row.append(mtb_p.summary_data[id])
				else:
					cur_row.append(None)

				if (b_in_c):
					cur_row.append(mtb_n.summary_data[id])
				else:
					cur_row.append(None)
			except:
				conf.error(f"ERROR: \tBAD ID FILTER: \"{id}\"")
				break

		ws.append(cur_row)

	conf.log("\tsetting column dimensions")
	dim_h["name_group"] = ColumnDimension(ws, min=1, max=2, width=5)
	dim_h["mtb_name"] = ColumnDimension(ws, min=3, max=3, width=30)
	dim_h["final_data"] = ColumnDimension(ws, min=(id_offset), max=id_offset + (conf.num_id_filters * 2), width=10)
	ws.column_dimensions = dim_h

	conf.log("creating data table sheet")
	ws = wb.create_sheet("Data Table")

	conf.log("\tadding headers")
	cur_row = []
	cur_row.append("Name")
	cur_row.append("Group")

	pop_i = []
	for i in range(0, len(mode_s.mtb)):
		if (mode_s.mtb[i].name in conf.exclusions):
			pop_i.append(i)

	num_pops = 0
	for i in pop_i:
		try:
			mode_s.mtb.pop(i - num_pops)
		except Exception as e:
			continue
		num_pops += 1

	for mtb in mode_s.mtb:
		cur_row.append(mtb.name)
	ws.append(cur_row)

	conf.log("\tadding metabolite data")

	ctr = 0
	group_ctr = 1
	group_col = []

	for g in conf.group:
		print(g)
		try:
			int(conf.group[g]) + 1
		except:
			continue
		for i in range(0, int(conf.group[g])):
			col = str(g) + " " + "{0:03d}".format(group_ctr)
			group_col.append(col)
			ctr += 1
			group_ctr += 1
		group_ctr = 1
		
	ctr = 0
	col_i = [
		int(get_col_letters_as_i(conf.final_col_range[0])),
		int(get_col_letters_as_i(conf.final_col_range[1])) ]
	for i in range(col_i[0], col_i[1]+1):
		cur_row = []
		name = "{0:03d}".format(ctr)
		name_display = "{0:03d}".format(ctr + 1)
		cur_row.append(name_display)
		try:
			cur_row.append(group_col[ctr])
		except:
			cur_row.append("")

		for mtb in mode_s.mtb:
			cur_row.append(mtb.group_data[name])
		ctr += 1

		ws.append(cur_row)

	conf.log("\tsetting column dimensions")
	dim_h = DimensionHolder(worksheet=ws)
	dim_h['all'] = ColumnDimension(ws, min=2, max=ws.max_column, width=20)
	ws.column_dimensions = dim_h

	wb.save(conf.stitch_filename)


def print_help():
	print(f"{"-h":<20} | print the help screen")
	print(f"{"-R":<20} | (GUI) specific command, this tells Ion Stitch to return a stitched data set.")
	print(f"{"-E":<20} | Export an ion stitched dataset as an excel sheet with two pages. The first is a Summary Table and the second is metaboanalyst ready.")
	print(f"{"--final_col_range":<20} | Takes the following two args (as column letters or numbers) and sets the column range that will be included in the dataset.")
	print(f"{"--exclude":<20} | Adds an exclusion rule for a specific compound, ignoring it in the final export.")
	print(f"{"--include":<20} | Deletes and exclusion rule for a specific compound.")
	print(f"{"-pos":<20} | Takes next arg as filename of positive ion mode.")
	print(f"{"-neg":<20} | Takes next arg as filename of negative ion mode.")
	print(f"{"-mf":<20} | Rearranges the order of a filter")
	print(f"{"-mg":<20} | Rearranges the order of a group")
	print(f"{"-rf":<20} | Removes every filter specified until another flag is detected.")
	print(f"{"-rg":<20} | Removes every group specified until another flag is detected.")
	print()
	print(f"{"-af(s)(A)":<20} | Adds a filter. The next arg is the filter id which appears in the summary sheet. If (A) is not provided then the following arg (ideally encapsulated with "") is the name of the column that filter will filter in. Otherwise it will take the next two args as the range of data to average.")
	print()
	print(f"{"-ag":<20} | Adds a group. The next arg is the group name. The arg after is the size of the group.")
	print(f"{"--prio_large":<20} | Takes a filter id and sets it's data priority to the larger of the two for stitching comparison.")
	print(f"{"--prio_large":<20} | Takes a filter id and sets it's data priority to the smaller of the two for stitching comparison.")
	print(f"{"--match_single":<20} | Takes a filter id and sets it's data collection style to be an exact match of the column header you are interested in comparing.")
	print(f"{"--match_average":<20} | Takes a filter id and sets it's data collection style to be an average every column between (inclusive) two column letters.")


def main(argv, ion_modes=None):
	conf = Conf()
	conf.load_conf(CONF_FILE)
	conf.read_args(argv)
	conf.save()
	print(conf.final_col_range[0])
	print(conf.final_col_range[1])

	if (ion_modes != None):
		if ("Export" in conf.mode):
			final_stitched_ion_mode = final_stitch(conf, ion_modes)
			conf.log(f"creating xlsx file: {conf.stitch_filename}")
			create_xl(ion_modes[0], ion_modes[1], final_stitched_ion_mode, conf)
			return

	if ("help" in conf.mode):
		print_help()
	elif (("Review" in conf.mode) or ("Export" in conf.mode)):
		conf.log("getting pos ion mode:", nl=False)
		pos_ion_mode = xl_data(conf.pos)
		pos_ion_mode.load_xl(conf)

		conf.log("getting neg ion mode:", nl=False)
		neg_ion_mode = xl_data(conf.neg)
		neg_ion_mode.load_xl(conf)

		conf.check_error()

		conf.log("stitching modes together")
		stitched_ion_mode = stitch_modes(conf, pos_ion_mode, neg_ion_mode)
		stitched_ion_mode.header = pos_ion_mode.header

		if ("Review" in conf.mode):
			return [pos_ion_mode, neg_ion_mode, stitched_ion_mode]
		if ("Export" in conf.mode):
			conf.log(f"creating xlsx file: {conf.stitch_filename}")
			create_xl(pos_ion_mode, neg_ion_mode, stitched_ion_mode, conf)


if __name__ == "__main__":
	main(sys.argv)
